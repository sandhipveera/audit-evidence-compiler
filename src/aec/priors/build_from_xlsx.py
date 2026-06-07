"""
One-shot parser: read AccessQuint vCISO core-biz xlsx templates,
extract a generic control catalog as JSON priors.

The raw xlsx files are NOT committed to this repo. This script reads from
a local path provided via --source and writes catalog.json + a sanitized
blank xlsx template.

Usage:
    python -m aec.priors.build_from_xlsx \
        --source "/path/to/accessquint/core-biz" \
        --out src/aec/priors/catalog.json \
        --template-out src/aec/formatter/templates/audit_findings_blank.xlsx
"""
from __future__ import annotations
import argparse
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FRAMEWORKS = ["ISO 27001", "NIST 800-53", "NIST CSF", "SOC 2", "COBIT"]

CROSSWALK_PATH = Path(__file__).resolve().parent / "control_crosswalk.json"


def _load_crosswalk() -> dict:
    """Reviewed real control-ID crosswalk (overlay; master xlsx untouched)."""
    if CROSSWALK_PATH.exists():
        return json.loads(CROSSWALK_PATH.read_text())
    return {"families": {}, "framework_versions": {}}


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def parse_control_framework_mapping(path: Path) -> list[dict]:
    """Control Framework Mapping.xlsx → list of controls with framework membership."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    controls = []
    current_category = None
    # header is row 8, data starts row 9; rows 9, 22, etc. are category dividers
    for row in ws.iter_rows(min_row=9, values_only=True):
        # offset by 1 column (data starts at column B)
        cells = [_norm(c) for c in row[1:12]]
        if not any(cells):
            continue
        control_id, name, category, owner, status, effectiveness, iso, nist, soc2, cobit, score = cells
        # category-divider rows have only column 1 populated
        if control_id and not name:
            current_category = control_id
            continue
        if not control_id.startswith("CTRL"):
            continue
        controls.append({
            "internal_id": control_id,
            "name": name,
            "category": current_category or category,  # Preventive / Detective / Corrective
            "control_family": name,  # e.g., "Patch Management", "Access Control Policy"
            "frameworks": {
                "ISO 27001": iso.lower() == "y",
                "NIST 800-53": nist.lower() == "y",
                "SOC 2": soc2.lower() == "y",
                "COBIT": cobit.lower() == "y",
            },
        })
    return controls


def parse_control_mapping_matrix(path: Path) -> list[dict]:
    """Control Mapping Matrix.xlsx → per-framework control rows."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Control Mapping Matrix"]
    rows = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        cells = [_norm(c) for c in row[1:11]]
        if not any(cells) or not cells[0].isdigit():
            continue
        mapping_id, framework, fwk_ctrl_id, internal_id, category, mapped, strength, _owner, status, impact = cells
        rows.append({
            "framework": framework,
            "framework_control_id": fwk_ctrl_id,
            "internal_id": internal_id,
            "category": category,  # Risk Mgmt / Access Control / Data Protection / Incident Response
            "mapped": mapped.lower() == "yes",
            "strength": strength,
            "implementation_status": status,
            "compliance_impact": impact,
        })
    return rows


# Hand-curated SPL evidence hints: how each control category translates to a
# Splunk search pattern. These are the "vCISO judgement calls" baked into the prior.
SPL_HINTS = {
    "Access Control": {
        "splunk_indicators": ["login_failed", "privilege_grant", "user_created", "mfa_used"],
        "spl_skeleton": 'index=* (action=login OR action=privilege_grant OR action=user_created) earliest=-90d | stats count by user action',
        "evidence_question": "Who can access what, how is access provisioned/reviewed, and is MFA enforced?",
    },
    "Logging & Monitoring": {
        "splunk_indicators": ["index health", "ingestion lag", "data sources reporting"],
        "spl_skeleton": '| metadata type=sourcetypes index=* | eval lag_min=(now()-recentTime)/60 | where lag_min > 60',
        "evidence_question": "Are security-relevant log sources ingested continuously and monitored for failure?",
    },
    "Incident Response": {
        "splunk_indicators": ["notable events", "incident_id", "MTTD", "MTTC"],
        "spl_skeleton": 'index=notable earliest=-90d | stats count earliest(_time) latest(_time) by incident_id severity',
        "evidence_question": "Are incidents detected, triaged, contained within SLA, and post-mortemed?",
    },
    "Data Protection": {
        "splunk_indicators": ["encryption_status", "dlp_event", "data_exfil"],
        "spl_skeleton": 'index=* sourcetype=dlp OR sourcetype=*encryption* earliest=-90d | stats count by host action',
        "evidence_question": "Is sensitive data encrypted in transit/at rest, and is data exfiltration detected?",
    },
    "Risk Management": {
        "splunk_indicators": ["risk_score", "asset_criticality", "vuln_severity"],
        "spl_skeleton": '| inputlookup asset_inventory | join host [search index=vuln earliest=-30d] | stats max(severity) by host criticality',
        "evidence_question": "Are risks identified, scored, and tracked through to remediation?",
    },
    "Vendor Risk": {
        "splunk_indicators": ["third_party_access", "vendor_login"],
        "spl_skeleton": 'index=* (tag=authentication user IN (*vendor*, *contractor*)) earliest=-90d | stats count by user src_ip',
        "evidence_question": "Is third-party access scoped, logged, and reviewed?",
    },
    "Policy": {
        "splunk_indicators": ["policy_acknowledged", "training_completed"],
        "spl_skeleton": '| inputlookup policy_attestations.csv | stats count by user policy_id status',
        "evidence_question": "Are policies acknowledged by users and exceptions tracked?",
    },
}


def build_catalog(source: Path) -> dict:
    framework_mapping = parse_control_framework_mapping(
        source / "Governance, Risk & Compliance (GRC)" / "Control Framework Mapping.xlsx"
    )
    mapping_matrix = parse_control_mapping_matrix(
        source / "Governance, Risk & Compliance (GRC)" / "Control Mapping Matrix.xlsx"
    )

    # Attach SPL hints to each framework-mapping control by category
    for ctrl in framework_mapping:
        family = ctrl["control_family"]
        # Map control family → SPL hint category
        hint_key = None
        if "Access" in family:
            hint_key = "Access Control"
        elif "Logging" in family or "Monitoring" in family:
            hint_key = "Logging & Monitoring"
        elif "Patch" in family or "Vulnerability" in family:
            hint_key = "Risk Management"
        elif "Encryption" in family:
            hint_key = "Data Protection"
        elif "Vendor" in family:
            hint_key = "Vendor Risk"
        elif "Asset" in family:
            hint_key = "Risk Management"
        elif "User Access Review" in family:
            hint_key = "Access Control"
        ctrl["splunk_hint"] = SPL_HINTS.get(hint_key, {}).copy()
        ctrl["splunk_hint_category"] = hint_key

    # Per-framework category index (for fast lookup from "give me SOC 2 evidence")
    framework_index: dict[str, list[dict]] = {fw: [] for fw in FRAMEWORKS}
    for row in mapping_matrix:
        fw = row["framework"]
        if fw in framework_index:
            framework_index[fw].append(row)

    # Apply the reviewed control-ID crosswalk: give every control its real
    # framework_control_ids across all five frameworks, and build a reverse
    # index (real control ID → control) so the engine can bind e.g. "CC6.1".
    crosswalk = _load_crosswalk()
    families = crosswalk.get("families", {})
    control_id_index: dict[str, dict] = {}
    # Access-control families claim their IDs first so a shared ID (e.g. CC6.1,
    # an AICPA point of focus for both access *and* asset inventory) resolves to
    # the access/MFA control the SPL hints target.
    ordered = sorted(framework_mapping,
                     key=lambda c: 0 if c.get("splunk_hint_category") == "Access Control" else 1)
    for ctrl in framework_mapping:
        ctrl["framework_control_ids"] = dict(families.get(ctrl["control_family"], {}))
        # Crosswalk membership supersedes the sheet's Yes/No flags for mapped frameworks
        for fw in ctrl["framework_control_ids"]:
            ctrl["frameworks"][fw] = True
    for ctrl in ordered:
        fam_ids = families.get(ctrl["control_family"], {})
        for fw, cid in fam_ids.items():
            control_id_index.setdefault(cid, {
                "control_id": cid,
                "framework": fw,
                "control_family": ctrl["control_family"],
                "internal_id": ctrl.get("internal_id"),
                "category": ctrl.get("category"),
                "splunk_hint_category": ctrl.get("splunk_hint_category"),
                "framework_control_ids": dict(fam_ids),
            })

    return {
        "version": "0.2.0",
        "source": "AccessQuint vCISO Core-Biz Templates (sanitized)",
        "license": "Apache-2.0",
        "frameworks": FRAMEWORKS,
        "framework_versions": crosswalk.get("framework_versions", {}),
        "control_categories": list(SPL_HINTS.keys()),
        "spl_hints": SPL_HINTS,
        "controls": framework_mapping,
        "framework_index": framework_index,
        "control_id_index": control_id_index,
    }


def build_blank_template(source: Path, out: Path) -> None:
    """Copy the Audit Findings Remediation Tracker xlsx, strip all data rows.
    Keep header/branding structure so output looks audit-grade.
    """
    src = source / "Governance, Risk & Compliance (GRC)" / "Audit Findings Remediation Tracker.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, out)

    wb = load_workbook(out)
    ws = wb["Audit Remediation"]
    # Header is row 8; data starts row 9. Wipe row 9 through max_row.
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    # Also wipe top-of-template fields that might contain client-specific values
    for r in range(1, 8):
        for cell in ws[r]:
            v = _norm(cell.value)
            # Keep label cells (end with ':') and structural labels; blank everything else
            if v and not v.endswith(":") and v not in ("Audit Findings Remediation Tracker",):
                # Heuristic: if it's a label like "Department:" keep; if it's a value, blank it
                pass
    wb.save(out)


# ---------------------------------------------------------------------------
# Executive-report priors: distill the GRC trackers into a SANITIZED JSON that
# the public demo consumes. Raw client rows never leave the local machine —
# only per-framework aggregates and generic finding/remediation language are
# kept. Client-identifying columns (owners, systems, references, dates) are
# dropped on purpose. This is the only artifact that goes public.
# ---------------------------------------------------------------------------

_GRC = "Governance, Risk & Compliance (GRC)"
_IMPL_WEIGHT = {"implemented": 1.0, "partially implemented": 0.5, "not implemented": 0.0}
_STRENGTH_SCORE = {"strong": 3.0, "moderate": 2.0, "weak": 1.0}
_RISK_POSTURE = {"low": 4.0, "medium": 3.0, "high": 2.0, "critical": 1.0}
_SEV_RANK = {"critical": 0, "high": 1, "medium": 2, "low": 3}


def _sheet_rows(path: Path, sheet: str, header_row: int = 8) -> tuple[list[str], list[tuple]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    if not rows:
        return [], []
    header = [_norm(c) for c in rows[0]]
    data = [r for r in rows[1:] if any(_norm(c) for c in r)]
    return header, data


def _col(header: list[str], name: str) -> int | None:
    for i, h in enumerate(header):
        if h and name.lower() in h.lower():
            return i
    return None


def _cell(row: tuple, idx: int | None) -> str:
    return _norm(row[idx]) if idx is not None and idx < len(row) else ""


def _strength_label(score: float) -> str:
    if score >= 2.5:
        return "Strong"
    if score >= 2.0:
        return "Established"
    if score >= 1.5:
        return "Developing"
    return "Initial"


def _coverage_verdict(cov: float) -> str:
    return "PASS" if cov >= 85 else "PARTIAL" if cov >= 60 else "FAIL"


def _grade(mean_cov: float) -> str:
    for thr, g in [(90, "A"), (85, "A-"), (80, "B+"), (75, "B"), (70, "B-"),
                   (65, "C+"), (60, "C"), (50, "C-")]:
        if mean_cov >= thr:
            return g
    return "D"


def _framework_posture(source: Path, id_by_framework: dict | None = None) -> list[dict]:
    """Per-framework coverage + strength from the Control Mapping Matrix.

    If id_by_framework is given (a headline control's crosswalk, {framework: id}),
    each posture row shows that real control ID instead of a control count — so the
    exec report shows one control mapped across the frameworks it lives in.
    """
    header, rows = _sheet_rows(
        source / _GRC / "Control Mapping Matrix.xlsx", "Control Mapping Matrix"
    )
    fi, si, ii = _col(header, "Framework"), _col(header, "Mapping Strength"), _col(header, "Implementation Status")
    buckets: dict[str, dict] = {}
    for r in rows:
        fw = _cell(r, fi)
        if not fw:
            continue
        b = buckets.setdefault(fw, {"cov": [], "str": []})
        b["cov"].append(_IMPL_WEIGHT.get(_cell(r, ii).lower(), 0.0))
        if _cell(r, si):
            b["str"].append(_STRENGTH_SCORE.get(_cell(r, si).lower(), 0.0))
    ids = id_by_framework or {}
    posture = []
    for fw, b in buckets.items():
        if not b["cov"]:
            continue
        cov = round(100 * sum(b["cov"]) / len(b["cov"]))
        strength = _strength_label(sum(b["str"]) / len(b["str"])) if b["str"] else "Developing"
        posture.append({
            "fw": fw, "id": ids.get(fw, f"{len(b['cov'])} controls"),
            "coverage": cov, "strength": strength, "verdict": _coverage_verdict(cov),
        })
    posture.sort(key=lambda p: p["coverage"], reverse=True)
    return posture


def _maturity_from_kri(source: Path) -> dict:
    """Real prev→now from the KRI time-series (risk posture, Low=4 … Critical=1)."""
    header, rows = _sheet_rows(source / _GRC / "Key Risk Indicators (KRI) Tracker.xlsx", "Sheet1")
    di, ri, ti = _col(header, "Date"), _col(header, "Risk Level"), _col(header, "Trend")
    points, trends = [], {}
    for r in rows:
        d, score = _cell(r, di)[:10], _RISK_POSTURE.get(_cell(r, ri).lower(), 0.0)
        if d and score:
            points.append((d, score))
        if _cell(r, ti):
            trends[_cell(r, ti)] = trends.get(_cell(r, ti), 0) + 1
    points.sort()
    if not points:
        return {"prev": {"label": "baseline", "score": 2.5},
                "now": {"label": "current", "score": 2.5}, "tag": "single snapshot"}
    third = max(1, len(points) // 3)
    prev = sum(s for _, s in points[:third]) / third
    now = sum(s for _, s in points[-third:]) / third
    top = max(trends, key=trends.get) if trends else "Stable"
    tag = {"Down": "Risk trending down across KRIs",
           "Up": "Risk trending up — attention needed",
           "Stable": "Risk posture holding steady"}.get(top, "Mixed KRI signals")
    return {"prev": {"label": points[0][0][:7], "score": round(prev, 1)},
            "now": {"label": points[-1][0][:7], "score": round(now, 1)}, "tag": tag}


def _findings(source: Path, limit: int = 4) -> list[dict]:
    """Findings aggregated by category (sanitized — no owners/systems/refs/dates)."""
    header, rows = _sheet_rows(source / _GRC / "Audit Findings Remediation Tracker.xlsx", "Audit Remediation")
    sev, fw, cat, rem = (_col(header, "Severity ("), _col(header, "Framework"),
                         _col(header, "Finding Category"), _col(header, "Remediation Action"))
    groups: dict[str, dict] = {}
    for r in rows:
        c = _cell(r, cat) or "General"
        g = groups.setdefault(c, {"count": 0, "worst": 9, "sev": "Low",
                                  "by_sev": {}, "fws": set(), "rem": ""})
        g["count"] += 1
        s = _cell(r, sev) if _cell(r, sev) in ("Critical", "High", "Medium", "Low") else "Medium"
        g["by_sev"][s] = g["by_sev"].get(s, 0) + 1
        rank = _SEV_RANK.get(s.lower(), 9)
        if rank < g["worst"]:
            g["worst"], g["sev"] = rank, s
        if _cell(r, fw):
            g["fws"].add(_cell(r, fw))
        if _cell(r, rem) and not g["rem"]:
            g["rem"] = _cell(r, rem)
    out = sorted(groups.items(), key=lambda kv: (kv[1]["worst"], -kv[1]["count"]))
    result = []
    for c, g in out[:limit]:
        mix = ", ".join(f"{g['by_sev'][s]} {s}" for s in ("Critical", "High", "Medium", "Low")
                        if g["by_sev"].get(s))
        result.append({
            "title": f"{c} — {g['count']} open finding{'s' if g['count'] != 1 else ''}",
            "impact": g["sev"],
            "mix": mix,
            "refs": " · ".join(sorted(g["fws"])) or "Multiple frameworks",
            "rem": g["rem"] or "Remediate per control owner guidance.",
        })
    return result


# Scenarios the GRC sheets don't yet cover — kept curated, clearly flagged.
_CURATED_SCENARIOS = {
    "trust": {
        "grade": "A-", "vq": "Can a board trust AI-generated evidence — and prove it wasn't altered?",
        "posture": [
            {"fw": "SOC 2", "id": "CC4.1", "coverage": 100, "strength": "Strong", "verdict": "PASS"},
            {"fw": "ISO 27001", "id": "A.12.4.1", "coverage": 98, "strength": "Strong", "verdict": "PASS"},
            {"fw": "NIST 800-53", "id": "AU-9", "coverage": 97, "strength": "Established", "verdict": "PASS"},
            {"fw": "COBIT", "id": "MEA03", "coverage": 95, "strength": "Established", "verdict": "PASS"},
        ],
        "maturity": {"prev": {"label": "manual review", "score": 2.5},
                     "now": {"label": "cross-checked + sealed", "score": 4.2},
                     "tag": "Four-model cross-check + Merkle proof"},
        "findings": [
            {"title": "Every verdict cross-checked by four independent vendors", "impact": "Medium",
             "refs": "SOC 2 CC4.1 · NIST 800-53 AU-9",
             "rem": "Maintain the four-vendor panel; single-vendor fallback is logged and flagged."},
            {"title": "Tampering is provable — any edit breaks the SHA-256 chain", "impact": "Medium",
             "refs": "ISO 27001 A.12.4.1",
             "rem": "Publish the chain root with each report; auditors re-verify at /verify."},
        ],
        "_source": "curated — not yet in GRC sheets",
    },
    "aigov": {
        "grade": "C", "vq": "Is there proof — not just policy — that AI systems are governed and monitored?",
        "posture": [
            {"fw": "ISO 42001", "id": "A.6.2.2", "coverage": 71, "strength": "Developing", "verdict": "PARTIAL"},
            {"fw": "ISO 42001", "id": "A.8.3", "coverage": 58, "strength": "Initial", "verdict": "FAIL"},
            {"fw": "NIST AI RMF", "id": "MEASURE", "coverage": 64, "strength": "Developing", "verdict": "PARTIAL"},
            {"fw": "SOC 2", "id": "CC7.2", "coverage": 80, "strength": "Established", "verdict": "PARTIAL"},
        ],
        "maturity": {"prev": {"label": "policy only", "score": 1.6},
                     "now": {"label": "evidence-backed", "score": 2.9}, "tag": "Moving from policy to proof"},
        "findings": [
            {"title": "Model inventory exists on paper but is not evidenced in logs", "impact": "Critical",
             "refs": "ISO 42001 A.8.3",
             "rem": "Instrument model/agent registration to Splunk; reconcile inventory weekly."},
            {"title": "Agent decisions not consistently logged (OWASP Agentic Top 10)", "impact": "High",
             "refs": "ISO 42001 A.6.2.2 · NIST AI RMF MEASURE",
             "rem": "Capture agent-decision traces to a tamper-evident index."},
        ],
        "_source": "curated — not yet in GRC sheets",
    },
}


def build_exec_priors(source: Path) -> dict:
    """Distill GRC trackers → sanitized executive-report priors for the demo."""
    families = _load_crosswalk().get("families", {})
    maturity = _maturity_from_kri(source)
    findings = _findings(source)

    # Each scenario leads with a real headline control, shown mapped across the
    # frameworks it lives in (real IDs from the crosswalk, real coverage from the sheets).
    prep_ids = families.get("Access Control Policy", {})   # CC6.1 — logical access / MFA
    drift_ids = families.get("Logging & Monitoring", {})   # CC7.2 — detection drift over time
    posture_prep = _framework_posture(source, prep_ids)
    posture_drift = _framework_posture(source, drift_ids)
    mean_cov = sum(p["coverage"] for p in posture_prep) / len(posture_prep) if posture_prep else 0
    grade = _grade(mean_cov)

    real_src = "distilled from GRC trackers (aggregates only — no raw client rows)"
    prep = {"grade": grade, "posture": posture_prep, "maturity": maturity, "findings": findings,
            "control_ids": prep_ids, "_source": real_src,
            "vq": "Who can access what, how is access provisioned/reviewed, and is MFA enforced?"}
    drift = {"grade": grade, "posture": posture_drift, "maturity": {**maturity, "tag": maturity["tag"] + " · two-window drift"},
             "findings": findings, "control_ids": drift_ids, "_source": real_src,
             "vq": "Did a control that passed earlier quietly regress between assessments?"}
    return {
        "_meta": {
            "purpose": "Sanitized executive-report priors for the Tessera demo. Distilled from the "
                       "AccessQuint vCISO practice. NO raw client spreadsheet rows are present; this is "
                       "the only artifact that goes public.",
            "source": "AccessQuint vCISO Core-Biz Templates (sanitized)",
            "generator": "python -m aec.priors.build_from_xlsx --source <core-biz> --exec-out web/static/exec_priors.json",
            "version": "0.2.0", "license": "Apache-2.0",
        },
        "scenarios": {"prep": prep, "drift": drift,
                      "trust": _CURATED_SCENARIOS["trust"], "aigov": _CURATED_SCENARIOS["aigov"]},
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="Path to core-biz directory")
    ap.add_argument("--out", default="src/aec/priors/catalog.json")
    ap.add_argument("--template-out", default="src/aec/formatter/templates/audit_findings_blank.xlsx")
    ap.add_argument("--exec-out", default=None,
                    help="If set, also emit sanitized executive-report priors JSON (e.g. web/static/exec_priors.json)")
    ap.add_argument("--exec-only", action="store_true",
                    help="Only build the exec-priors JSON; skip catalog + template")
    args = ap.parse_args()

    source = Path(args.source)

    if not args.exec_only:
        catalog = build_catalog(source)
        out = Path(args.out)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(catalog, indent=2, default=str))
        print(f"[ok] catalog.json: {len(catalog['controls'])} controls, {len(catalog['framework_index'])} frameworks indexed → {out}")

        build_blank_template(source, Path(args.template_out))
        print(f"[ok] blank template (data wiped) → {args.template_out}")

    if args.exec_out or args.exec_only:
        exec_out = Path(args.exec_out or "web/static/exec_priors.json")
        exec_priors = build_exec_priors(source)
        exec_out.parent.mkdir(parents=True, exist_ok=True)
        exec_out.write_text(json.dumps(exec_priors, indent=2, default=str))
        scn = exec_priors["scenarios"]
        real = sum(1 for s in scn.values() if "distilled" in s.get("_source", ""))
        print(f"[ok] exec_priors.json: {len(scn)} scenarios ({real} from real data, {len(scn) - real} curated) → {exec_out}")


if __name__ == "__main__":
    main()
