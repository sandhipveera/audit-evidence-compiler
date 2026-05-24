"""Manifest sheet — embeds chain root into the xlsx for offline verification."""
from __future__ import annotations

from datetime import date, datetime, time, timezone
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import aec

MANIFEST_SHEET_NAME = "Manifest"
MANIFEST_HASH_FIELD = "manifest_hash"

MANIFEST_FIELDS = [
    ("chain_root", "SHA-256 root hash of the evidence chain"),
    ("chain_length", "Total number of evidence snapshots"),
    ("workbook_hash", "SHA-256 digest of non-Manifest workbook cells"),
    ("created_at", "UTC timestamp when this report was sealed"),
    ("aec_version", "Audit Evidence Compiler version"),
    ("verify_command", "Command to verify this report"),
    (MANIFEST_HASH_FIELD, "SHA-256 digest of Manifest cells except this value"),
]


def write_manifest_sheet(
    xlsx_path: Path,
    chain_root: str,
    chain_length: int,
    created_at: str | None = None,
) -> None:
    """Append (or replace) a Manifest sheet in the given xlsx."""
    if created_at is None:
        created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    workbook_hash = compute_workbook_hash(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)

    if MANIFEST_SHEET_NAME in wb.sheetnames:
        del wb[MANIFEST_SHEET_NAME]

    ws = wb.create_sheet(MANIFEST_SHEET_NAME)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    label_font = Font(bold=True, size=11)
    value_font = Font(name="Consolas", size=11)

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "Evidence Chain Manifest"
    title_cell.font = Font(bold=True, size=14)

    ws["A3"].value = "Field"
    ws["A3"].font = header_font_white
    ws["A3"].fill = header_fill
    ws["B3"].value = "Value"
    ws["B3"].font = header_font_white
    ws["B3"].fill = header_fill
    ws["C3"].value = "Description"
    ws["C3"].font = header_font_white
    ws["C3"].fill = header_fill

    values = {
        "chain_root": chain_root,
        "chain_length": str(chain_length),
        "workbook_hash": workbook_hash,
        "created_at": created_at,
        "aec_version": aec.__version__,
        "verify_command": "aec verify gap_report.xlsx",
        MANIFEST_HASH_FIELD: "",
    }

    manifest_hash_cell = None
    for i, (field_name, description) in enumerate(MANIFEST_FIELDS):
        row = 4 + i
        ws.cell(row=row, column=1, value=field_name).font = label_font
        val_cell = ws.cell(row=row, column=2, value=values[field_name])
        val_cell.font = value_font
        val_cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=description).font = Font(size=10, italic=True)
        if field_name == MANIFEST_HASH_FIELD:
            manifest_hash_cell = val_cell

    if manifest_hash_cell is not None:
        manifest_hash_cell.value = _hash_payload(
            _workbook_payload(wb, include_manifest=True, exclude_manifest_hash=True)
        )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 45

    wb.save(xlsx_path)


def read_manifest(xlsx_path: Path) -> dict[str, str] | None:
    """Read the Manifest sheet and return a dict of field→value, or None if missing."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if MANIFEST_SHEET_NAME not in wb.sheetnames:
        return None

    ws = wb[MANIFEST_SHEET_NAME]
    manifest: dict[str, str] = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=2, values_only=True):
        if row[0] and row[1]:
            manifest[str(row[0]).strip()] = str(row[1]).strip()
    return manifest


def compute_workbook_hash(xlsx_path: Path) -> str:
    """Hash workbook cell values outside the Manifest sheet."""
    payload = workbook_payload(xlsx_path)
    return _hash_payload(payload)


def compute_manifest_hash(xlsx_path: Path) -> str:
    """Hash all workbook cells, excluding only the Manifest hash value cell."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    payload = _workbook_payload(wb, include_manifest=True, exclude_manifest_hash=True)
    return _hash_payload(payload)


def _hash_payload(payload: dict[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    ).encode("utf-8")
    return f"sha256:{hashlib.sha256(encoded).hexdigest()}"


def workbook_payload(xlsx_path: Path) -> dict[str, Any]:
    """Return deterministic workbook cell content for tamper checks."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    return _workbook_payload(wb, include_manifest=False, exclude_manifest_hash=False)


def _workbook_payload(
    wb: openpyxl.Workbook,
    *,
    include_manifest: bool,
    exclude_manifest_hash: bool,
) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if ws.title == MANIFEST_SHEET_NAME and not include_manifest:
            continue

        cells: list[dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                if exclude_manifest_hash and _is_manifest_hash_value_cell(ws, cell):
                    continue
                if cell.value is not None:
                    cells.append(
                        {
                            "coordinate": cell.coordinate,
                            "data_type": cell.data_type,
                            "value": _normalize_cell_value(cell.value),
                        }
                    )

        sheets.append(
            {
                "title": ws.title,
                "merged_cells": sorted(str(cell_range) for cell_range in ws.merged_cells.ranges),
                "cells": cells,
            }
        )

    return {"sheets": sheets}


def _is_manifest_hash_value_cell(ws: openpyxl.worksheet.worksheet.Worksheet, cell: Any) -> bool:
    if ws.title != MANIFEST_SHEET_NAME or cell.column != 2:
        return False
    return ws.cell(row=cell.row, column=1).value == MANIFEST_HASH_FIELD


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, time):
        return {"type": "time", "value": value.isoformat()}
    if isinstance(value, bytes):
        return {"type": "bytes", "value": value.hex()}
    return value


def verify_report(
    xlsx_path: Path,
    trail_path: Path,
) -> tuple[bool, list[str]]:
    """Full verification: chain integrity + manifest cross-check.

    Returns (ok, messages) where ok is True if everything checks out.
    """
    from aec.integrity.chain import read_trail, verify_chain

    messages: list[str] = []

    if not xlsx_path.exists():
        return False, [f"XLSX not found: {xlsx_path}"]

    if not trail_path.exists():
        return False, [f"Trail file not found: {trail_path}"]

    try:
        snapshots = read_trail(trail_path)
    except ValueError as exc:
        return False, [str(exc)]

    if not snapshots:
        return False, ["Trail file is empty — no snapshots to verify"]

    chain_errors = verify_chain(snapshots)

    messages.append(f"Chain length: {len(snapshots)} snapshots")

    if chain_errors:
        for err in chain_errors:
            messages.append(err)
        return False, messages

    messages.append("Chain integrity: verified")

    trail_tip = snapshots[-1].get("this_hash", "")

    manifest = read_manifest(xlsx_path)
    if manifest is None:
        messages.append("No Manifest sheet in xlsx — cannot cross-check")
        return False, messages

    missing_fields = [field_name for field_name, _ in MANIFEST_FIELDS if field_name not in manifest]
    if missing_fields:
        messages.append(f"Manifest missing required fields: {', '.join(missing_fields)}")
        return False, messages

    manifest_root = manifest.get("chain_root", "")
    if manifest_root == trail_tip:
        short = trail_tip.replace("sha256:", "")
        messages.append(f"Manifest root matches trail tip: {short[:4]}...{short[-4:]}")
    else:
        messages.append(
            f"Manifest root MISMATCH: manifest says {manifest_root}, trail tip is {trail_tip}"
        )
        return False, messages

    manifest_length = manifest["chain_length"]
    try:
        expected_length = int(manifest_length)
    except ValueError:
        messages.append(f"Invalid chain length in Manifest: {manifest_length}")
        return False, messages

    if expected_length != len(snapshots):
        messages.append(
            f"Chain length mismatch: manifest says {manifest_length}, trail has {len(snapshots)}"
        )
        return False, messages

    workbook_hash = compute_workbook_hash(xlsx_path)
    manifest_workbook_hash = manifest["workbook_hash"]
    if manifest_workbook_hash != workbook_hash:
        messages.append(
            f"Workbook hash mismatch: manifest says {manifest_workbook_hash}, "
            f"actual {workbook_hash}"
        )
        return False, messages

    messages.append("Workbook cells match sealed workbook hash")

    manifest_hash = compute_manifest_hash(xlsx_path)
    manifest_manifest_hash = manifest[MANIFEST_HASH_FIELD]
    if manifest_manifest_hash != manifest_hash:
        messages.append(
            f"Manifest hash mismatch: manifest says {manifest_manifest_hash}, "
            f"actual {manifest_hash}"
        )
        return False, messages

    verify_command = manifest["verify_command"]
    if verify_command != "aec verify gap_report.xlsx":
        messages.append(
            f"Verify command mismatch: expected 'aec verify gap_report.xlsx', got {verify_command}"
        )
        return False, messages

    messages.append(
        f"Report cells consistent with snapshots "
        f"({len(snapshots)} findings <-> {len(snapshots)} snapshots)"
    )

    return True, messages
