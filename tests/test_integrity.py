"""Tests for Merkle-chained evidence trail — golden chain + tamper scenarios."""
from __future__ import annotations

import copy
import json
import uuid
from pathlib import Path

from aec.integrity.chain import (
    GENESIS,
    canonical_json,
    chain_snapshots,
    compute_snapshot_hash,
    read_trail,
    verify_chain,
    write_trail,
)
from aec.integrity.manifest import (
    compute_manifest_hash,
    compute_workbook_hash,
    read_manifest,
    verify_report,
    write_manifest_sheet,
)


def _make_snapshot(control_id: str = "CC6.1.c", row_count: int = 247) -> dict:
    return {
        "snapshot_id": str(uuid.uuid4()),
        "control_id": control_id,
        "spl_executed": f'index=* action=login earliest=-90d | stats count by user | where count > {row_count}',
        "row_count": row_count,
        "panel_verdict": "FAIL",
        "panel_transcript_hash": "sha256:abc123",
        "timestamp": "2026-06-01T14:23:00Z",
    }


def _make_chain(n: int = 5) -> list[dict]:
    controls = ["CC6.1.a", "CC6.1.b", "CC6.1.c", "CC7.2.a", "CC7.2.b"]
    snapshots = [_make_snapshot(controls[i % len(controls)], row_count=100 + i) for i in range(n)]
    return chain_snapshots(snapshots)


class TestCanonicalJson:
    def test_deterministic(self):
        snap = _make_snapshot()
        assert canonical_json(snap) == canonical_json(snap)

    def test_sorted_keys(self):
        raw = canonical_json(_make_snapshot()).decode("utf-8")
        parsed = json.loads(raw)
        assert list(parsed.keys()) == sorted(parsed.keys())

    def test_no_extraneous_whitespace(self):
        raw = canonical_json(_make_snapshot()).decode("utf-8")
        assert "\n" not in raw
        assert ": " not in raw
        assert ", " not in raw

    def test_excludes_hash_fields(self):
        snap = _make_snapshot()
        snap["prev_hash"] = "sha256:should_be_excluded"
        snap["this_hash"] = "sha256:also_excluded"
        raw = canonical_json(snap).decode("utf-8")
        assert "prev_hash" not in raw
        assert "this_hash" not in raw

    def test_utf8_encoding(self):
        snap = _make_snapshot()
        snap["spl_executed"] = 'index=* user="café" | stats count'
        raw = canonical_json(snap)
        assert "café".encode("utf-8") in raw


class TestComputeSnapshotHash:
    def test_returns_sha256_prefix(self):
        h = compute_snapshot_hash(_make_snapshot())
        assert h.startswith("sha256:")
        assert len(h) == len("sha256:") + 64

    def test_same_input_same_hash(self):
        snap = _make_snapshot()
        snap["snapshot_id"] = "fixed-id"
        h1 = compute_snapshot_hash(snap)
        h2 = compute_snapshot_hash(copy.deepcopy(snap))
        assert h1 == h2

    def test_different_input_different_hash(self):
        s1 = _make_snapshot()
        s1["snapshot_id"] = "id-1"
        s2 = _make_snapshot()
        s2["snapshot_id"] = "id-2"
        assert compute_snapshot_hash(s1) != compute_snapshot_hash(s2)

    def test_hash_ignores_prev_and_this_hash(self):
        snap = _make_snapshot()
        snap["snapshot_id"] = "fixed-id"
        h_bare = compute_snapshot_hash(snap)
        snap["prev_hash"] = "sha256:anything"
        snap["this_hash"] = "sha256:something_else"
        h_with = compute_snapshot_hash(snap)
        assert h_bare == h_with


class TestChainSnapshots:
    def test_first_snapshot_prev_hash_is_genesis(self):
        chain = _make_chain(1)
        assert chain[0]["prev_hash"] == GENESIS

    def test_chain_links(self):
        chain = _make_chain(5)
        for i in range(1, len(chain)):
            assert chain[i]["prev_hash"] == chain[i - 1]["this_hash"]

    def test_all_hashes_present(self):
        chain = _make_chain(3)
        for snap in chain:
            assert "prev_hash" in snap
            assert "this_hash" in snap
            assert snap["this_hash"].startswith("sha256:")

    def test_hashes_are_valid(self):
        chain = _make_chain(5)
        for snap in chain:
            expected = compute_snapshot_hash(snap)
            assert snap["this_hash"] == expected


class TestVerifyChain:
    def test_valid_chain_returns_no_errors(self):
        chain = _make_chain(10)
        errors = verify_chain(chain)
        assert errors == []

    def test_empty_chain_returns_no_errors(self):
        assert verify_chain([]) == []

    def test_single_snapshot(self):
        chain = _make_chain(1)
        assert verify_chain(chain) == []

    def test_tampered_row_count_detected(self):
        chain = _make_chain(5)
        chain[2]["row_count"] = 9999
        errors = verify_chain(chain)
        assert len(errors) >= 1
        assert "hash mismatch" in errors[0]
        assert "#3" in errors[0]

    def test_tampered_verdict_detected(self):
        chain = _make_chain(5)
        chain[0]["panel_verdict"] = "PASS"
        errors = verify_chain(chain)
        assert len(errors) >= 1

    def test_swapped_snapshots_detected(self):
        chain = _make_chain(5)
        chain[1], chain[2] = chain[2], chain[1]
        errors = verify_chain(chain)
        assert len(errors) >= 1

    def test_deleted_snapshot_detected(self):
        chain = _make_chain(5)
        del chain[2]
        errors = verify_chain(chain)
        assert len(errors) >= 1

    def test_inserted_snapshot_detected(self):
        chain = _make_chain(5)
        fake = _make_snapshot("FAKE.1")
        fake["prev_hash"] = chain[1]["this_hash"]
        fake["this_hash"] = compute_snapshot_hash(fake)
        chain.insert(2, fake)
        errors = verify_chain(chain)
        assert len(errors) >= 1

    def test_tampered_prev_hash_detected(self):
        chain = _make_chain(3)
        chain[1]["prev_hash"] = "sha256:0000000000000000000000000000000000000000000000000000000000000000"
        errors = verify_chain(chain)
        assert len(errors) >= 1
        assert "prev_hash mismatch" in errors[0]


class TestTrailIO:
    def test_roundtrip(self, tmp_path: Path):
        chain = _make_chain(5)
        trail_path = tmp_path / "audit_trail.jsonl"
        write_trail(trail_path, chain)
        loaded = read_trail(trail_path)
        assert len(loaded) == 5
        for orig, loaded_snap in zip(chain, loaded):
            assert orig["this_hash"] == loaded_snap["this_hash"]
            assert orig["prev_hash"] == loaded_snap["prev_hash"]

    def test_verify_after_roundtrip(self, tmp_path: Path):
        chain = _make_chain(10)
        trail_path = tmp_path / "audit_trail.jsonl"
        write_trail(trail_path, chain)
        loaded = read_trail(trail_path)
        errors = verify_chain(loaded)
        assert errors == []


class TestManifestSheet:
    def _make_xlsx(self, tmp_path: Path) -> Path:
        import openpyxl

        xlsx_path = tmp_path / "gap_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Findings"
        ws.append(["Control ID", "Verdict", "Details"])
        wb.save(xlsx_path)
        return xlsx_path

    def test_write_and_read(self, tmp_path: Path):
        xlsx = self._make_xlsx(tmp_path)
        write_manifest_sheet(xlsx, "sha256:deadbeef", 12)
        manifest = read_manifest(xlsx)
        assert manifest is not None
        assert manifest["chain_root"] == "sha256:deadbeef"
        assert manifest["chain_length"] == "12"
        assert manifest["mcp_server"] == "null"
        assert manifest["workbook_hash"].startswith("sha256:")
        assert manifest["manifest_hash"].startswith("sha256:")
        assert manifest["manifest_hash"] == compute_manifest_hash(xlsx)
        assert "aec_version" in manifest

    def test_manifest_records_mcp_server(self, tmp_path: Path):
        xlsx = self._make_xlsx(tmp_path)
        write_manifest_sheet(
            xlsx,
            "sha256:deadbeef",
            12,
            mcp_server="splunk-official-0.3.2",
        )
        manifest = read_manifest(xlsx)
        assert manifest is not None
        assert manifest["mcp_server"] == "splunk-official-0.3.2"

    def test_overwrite_manifest(self, tmp_path: Path):
        xlsx = self._make_xlsx(tmp_path)
        write_manifest_sheet(xlsx, "sha256:first", 5)
        write_manifest_sheet(xlsx, "sha256:second", 10)
        manifest = read_manifest(xlsx)
        assert manifest["chain_root"] == "sha256:second"
        assert manifest["chain_length"] == "10"

    def test_workbook_hash_ignores_manifest_sheet(self, tmp_path: Path):
        xlsx = self._make_xlsx(tmp_path)
        write_manifest_sheet(xlsx, "sha256:first", 5)
        first_hash = compute_workbook_hash(xlsx)
        write_manifest_sheet(xlsx, "sha256:second", 10)
        second_hash = compute_workbook_hash(xlsx)
        assert first_hash == second_hash

    def test_no_manifest_returns_none(self, tmp_path: Path):
        xlsx = self._make_xlsx(tmp_path)
        assert read_manifest(xlsx) is None

    def test_original_sheets_preserved(self, tmp_path: Path):
        import openpyxl

        xlsx = self._make_xlsx(tmp_path)
        write_manifest_sheet(xlsx, "sha256:test", 1)
        wb = openpyxl.load_workbook(xlsx)
        assert "Findings" in wb.sheetnames
        assert "Manifest" in wb.sheetnames


class TestVerifyReport:
    def _setup(self, tmp_path: Path, n: int = 5, tamper_fn=None):
        import openpyxl

        chain = _make_chain(n)
        trail_path = tmp_path / "audit_trail.jsonl"
        write_trail(trail_path, chain)

        xlsx_path = tmp_path / "gap_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Findings"
        for snap in chain:
            ws.append([snap["control_id"], snap["panel_verdict"], snap["row_count"]])
        wb.save(xlsx_path)

        chain_root = chain[-1]["this_hash"]
        write_manifest_sheet(xlsx_path, chain_root, len(chain))

        if tamper_fn:
            tamper_fn(trail_path, xlsx_path, chain)

        return xlsx_path, trail_path

    def test_intact_report_passes(self, tmp_path: Path):
        xlsx, trail = self._setup(tmp_path)
        ok, messages = verify_report(xlsx, trail)
        assert ok
        assert any("verified" in m for m in messages)
        assert any("matches" in m.lower() for m in messages)

    def test_tampered_trail_fails(self, tmp_path: Path):
        def tamper(trail_path, xlsx_path, chain):
            snapshots = read_trail(trail_path)
            snapshots[2]["row_count"] = 9999
            write_trail(trail_path, snapshots)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("mismatch" in m.lower() for m in messages)

    def test_tampered_xlsx_manifest_fails(self, tmp_path: Path):
        def tamper(trail_path, xlsx_path, chain):
            write_manifest_sheet(xlsx_path, "sha256:tampered_root_hash", len(chain))

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("mismatch" in m.upper() or "MISMATCH" in m for m in messages)

    def test_tampered_xlsx_data_cell_fails(self, tmp_path: Path):
        import openpyxl

        def tamper(trail_path, xlsx_path, chain):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["Findings"]
            ws["C3"] = 9999
            wb.save(xlsx_path)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("workbook hash mismatch" in m.lower() for m in messages)

    def test_tampered_xlsx_empty_cell_fails(self, tmp_path: Path):
        import openpyxl

        def tamper(trail_path, xlsx_path, chain):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["Findings"]
            ws["E20"] = "post-collection note"
            wb.save(xlsx_path)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("workbook hash mismatch" in m.lower() for m in messages)

    def test_tampered_manifest_description_cell_fails(self, tmp_path: Path):
        import openpyxl

        def tamper(trail_path, xlsx_path, chain):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["Manifest"]
            ws["C4"] = "edited after collection"
            wb.save(xlsx_path)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("manifest hash mismatch" in m.lower() for m in messages)

    def test_missing_trail_fails(self, tmp_path: Path):
        import openpyxl

        xlsx_path = tmp_path / "gap_report.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        ok, messages = verify_report(xlsx_path, tmp_path / "nonexistent.jsonl")
        assert not ok

    def test_missing_xlsx_fails(self, tmp_path: Path):
        chain = _make_chain(3)
        trail_path = tmp_path / "audit_trail.jsonl"
        write_trail(trail_path, chain)

        ok, messages = verify_report(tmp_path / "missing.xlsx", trail_path)
        assert not ok
        assert any("XLSX not found" in m for m in messages)

    def test_missing_manifest_fails(self, tmp_path: Path):
        import openpyxl

        chain = _make_chain(3)
        trail_path = tmp_path / "audit_trail.jsonl"
        write_trail(trail_path, chain)

        xlsx_path = tmp_path / "gap_report.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)

        ok, messages = verify_report(xlsx_path, trail_path)
        assert not ok
        assert any("No Manifest" in m for m in messages)

    def test_chain_length_mismatch_fails(self, tmp_path: Path):
        def tamper(trail_path, xlsx_path, chain):
            write_manifest_sheet(xlsx_path, chain[-1]["this_hash"], 999)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("length mismatch" in m.lower() for m in messages)

    def test_invalid_chain_length_fails_without_crashing(self, tmp_path: Path):
        import openpyxl

        def tamper(trail_path, xlsx_path, chain):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["Manifest"]
            for row in range(4, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == "chain_length":
                    ws.cell(row=row, column=2).value = "not-a-number"
                    break
            wb.save(xlsx_path)

        xlsx, trail = self._setup(tmp_path, tamper_fn=tamper)
        ok, messages = verify_report(xlsx, trail)
        assert not ok
        assert any("invalid chain length" in m.lower() for m in messages)

    def test_invalid_json_trail_fails_without_crashing(self, tmp_path: Path):
        import openpyxl

        xlsx_path = tmp_path / "gap_report.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        trail_path = tmp_path / "audit_trail.jsonl"
        trail_path.write_text('{"snapshot_id": "ok"}\n{bad json}\n', encoding="utf-8")

        ok, messages = verify_report(xlsx_path, trail_path)
        assert not ok
        assert any("Invalid JSON" in m for m in messages)

    def test_large_chain_performance(self, tmp_path: Path):
        """100-snapshot chain should verify in under 2 seconds."""
        import time

        xlsx, trail = self._setup(tmp_path, n=100)
        start = time.monotonic()
        ok, messages = verify_report(xlsx, trail)
        elapsed = time.monotonic() - start
        assert ok
        assert elapsed < 2.0, f"Verification took {elapsed:.2f}s (limit: 2.0s)"
