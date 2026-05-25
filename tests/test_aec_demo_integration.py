"""Integration test — aec_demo produces all 4 artifacts and aec verify passes."""
from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import patch

import pytest

from aec.agent.models import Critique, PanelResult
from aec.integrity.manifest import verify_report


def _fake_panel_result() -> PanelResult:
    return PanelResult(
        critiques=[
            Critique(
                persona="auditor",
                model="claude-sonnet-4",
                transport="anthropic-cli",
                verdict="PARTIAL",
                confidence=0.80,
                rationale="MFA coverage at 83% is below the 95% threshold.",
            ),
            Critique(
                persona="engineer",
                model="gpt-5",
                transport="openai-cli",
                verdict="PARTIAL",
                confidence=0.75,
                rationale="12 service accounts bypass MFA entirely.",
            ),
            Critique(
                persona="adversary",
                model="gemini-2.5-pro",
                transport="gemini-cli",
                verdict="FAIL",
                confidence=0.90,
                rationale="Service account MFA bypass is a critical gap.",
                concerns=["svc_deploy bypasses MFA"],
                recommended_additional_searches=[
                    "index=auth mfa_status=bypassed | stats count by user"
                ],
            ),
        ],
        final_verdict="FAIL",
        consensus_method="lowest_of_three",
        transcript="Auditor: PARTIAL\nEngineer: PARTIAL\nAdversary: FAIL\nConsensus: FAIL",
    )


@pytest.fixture
def demo_output(tmp_path, monkeypatch):
    """Run aec_demo with mocked panel and capture output files."""
    monkeypatch.chdir(tmp_path)
    (tmp_path / "out").mkdir()

    fake_result = _fake_panel_result()

    async def fake_run_panel(**kwargs):
        return fake_result

    def fake_format_transcript(result, cid, snap_name):
        return f"# Transcript\n{result.transcript}\n"

    with patch("aec.agent.panel.run_panel", new=fake_run_panel), \
         patch("aec.agent.panel._format_transcript_file", new=fake_format_transcript):
        import asyncio
        from cli.aec_demo import _run
        import argparse

        args = argparse.Namespace(
            sample="soc2-cc61",
            control=None,
            window="30d",
            no_llm=False,
        )
        asyncio.run(_run(args))

    out = tmp_path / "out"
    return out


def _find_file(directory: Path, prefix: str, suffix: str) -> Path:
    matches = list(directory.glob(f"{prefix}*{suffix}"))
    assert matches, f"No {prefix}*{suffix} in {directory}"
    return matches[0]


class TestDemoProducesAllArtifacts:
    def test_four_artifacts_exist(self, demo_output):
        memo = _find_file(demo_output, "audit_memo_", ".md")
        transcript = _find_file(demo_output, "transcript_", ".md")
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        xlsx = _find_file(demo_output, "gap_report_", ".xlsx")

        assert memo.stat().st_size > 0
        assert transcript.stat().st_size > 0
        assert trail.stat().st_size > 0
        assert xlsx.stat().st_size > 0

    def test_trail_has_four_snapshots(self, demo_output):
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        snapshots = [json.loads(line) for line in trail.read_text().strip().splitlines()]
        assert len(snapshots) == 4
        personas = [s["persona"] for s in snapshots]
        assert "auditor" in personas
        assert "engineer" in personas
        assert "adversary" in personas
        assert "consensus" in personas

    def test_trail_chain_is_valid(self, demo_output):
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        snapshots = [json.loads(line) for line in trail.read_text().strip().splitlines()]
        assert snapshots[0]["prev_hash"] == "sha256:GENESIS"
        for i in range(1, len(snapshots)):
            assert snapshots[i]["prev_hash"] == snapshots[i - 1]["this_hash"]

    def test_verify_passes(self, demo_output):
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        xlsx = _find_file(demo_output, "gap_report_", ".xlsx")
        ok, messages = verify_report(xlsx, trail)
        assert ok, f"Verify failed: {messages}"

    def test_tampered_xlsx_fails_verify(self, demo_output):
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        xlsx = _find_file(demo_output, "gap_report_", ".xlsx")

        import openpyxl
        tampered = demo_output / "gap_report_tampered.xlsx"
        wb = openpyxl.load_workbook(xlsx)
        wb.active["B9"] = "HACKED"
        wb.save(tampered)

        ok, messages = verify_report(tampered, trail)
        assert not ok, f"Tampered report should fail verify but got: {messages}"

    def test_tampered_trail_fails_verify(self, demo_output):
        trail = _find_file(demo_output, "audit_trail_", ".jsonl")
        xlsx = _find_file(demo_output, "gap_report_", ".xlsx")

        lines = trail.read_text().strip().splitlines()
        corrupted = json.loads(lines[0])
        corrupted["panel_verdict"] = "HACKED"
        lines[0] = json.dumps(corrupted, sort_keys=True)

        tampered_trail = demo_output / "audit_trail_tampered.jsonl"
        tampered_trail.write_text("\n".join(lines) + "\n")

        ok, messages = verify_report(xlsx, tampered_trail)
        assert not ok, f"Tampered trail should fail verify but got: {messages}"

    def test_gap_finding_has_high_severity_for_fail(self, demo_output):
        import openpyxl
        xlsx = _find_file(demo_output, "gap_report_", ".xlsx")
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["Audit Remediation"]
        severity_cell = ws.cell(row=9, column=8).value  # column H = severity
        assert severity_cell == "High"
