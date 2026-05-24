"""Tests for audit findings xlsx formatter."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from aec.formatter.audit_findings import (
    DATA_START_COL,
    DATA_START_ROW,
    SEVERITY_SCORES,
    SHEET_NAME,
    GapFinding,
    load_findings_from_json,
    write_findings,
)


def _make_finding(**overrides) -> GapFinding:
    defaults = {
        "finding_id": "FIND-001",
        "audit_type": "Internal",
        "framework": "SOC 2",
        "audit_reference": "CC6.1",
        "finding_description": "MFA not enforced for admin accounts",
        "finding_category": "Access Control",
        "severity": "High",
        "root_cause": "Policy gap in IAM configuration",
        "affected_system": "IAM / SSO",
        "risk_owner": "CISO",
        "remediation_action": "Enable MFA for all admin accounts",
        "remediation_owner": "IT Security",
        "current_status": "Open",
        "evidence_reference": "audit_trail.jsonl#snap-001",
        "comments": "",
    }
    defaults.update(overrides)
    return GapFinding(**defaults)


def _make_fixtures() -> list[GapFinding]:
    return [
        _make_finding(
            finding_id="FIND-001",
            severity="Critical",
            framework="SOC 2",
            audit_reference="CC6.1",
            finding_description="No MFA on privileged accounts",
            finding_category="Access Control",
            root_cause="IAM policy not enforcing MFA",
            affected_system="IAM / SSO",
            remediation_action="Enforce MFA for all privileged users",
        ),
        _make_finding(
            finding_id="FIND-002",
            severity="High",
            framework="ISO 27001",
            audit_reference="A.12.4.1",
            finding_description="Audit logs not retained for 12 months",
            finding_category="Logging & Monitoring",
            root_cause="Retention policy set to 30 days",
            affected_system="Splunk / SIEM",
            remediation_action="Extend log retention to 12 months",
        ),
        _make_finding(
            finding_id="FIND-003",
            severity="Medium",
            framework="NIST CSF",
            audit_reference="PR.AC-1",
            finding_description="Stale service accounts not deprovisioned",
            finding_category="Access Control",
            root_cause="No automated deprovisioning workflow",
            affected_system="Active Directory",
            remediation_action="Implement 90-day access review automation",
        ),
        _make_finding(
            finding_id="FIND-004",
            severity="Low",
            framework="SOC 2",
            audit_reference="CC7.2",
            finding_description="Incident response plan not tested annually",
            finding_category="Incident Response",
            root_cause="No scheduled tabletop exercises",
            affected_system="IR process",
            remediation_action="Schedule annual tabletop exercise",
        ),
    ]


REF_DATE = date(2026, 5, 24)


class TestGapFindingModel:
    def test_severity_score_mapping(self):
        for label, expected in SEVERITY_SCORES.items():
            f = _make_finding(severity=label.capitalize())
            assert f.severity_score == expected

    def test_severity_score_case_insensitive(self):
        assert _make_finding(severity="HIGH").severity_score == 3
        assert _make_finding(severity="low").severity_score == 1

    def test_severity_score_unknown_defaults_to_2(self):
        assert _make_finding(severity="Unknown").severity_score == 2

    def test_target_closure_critical(self):
        f = _make_finding(severity="Critical")
        assert f.target_closure_date(REF_DATE) == REF_DATE + __import__("datetime").timedelta(days=90)

    def test_target_closure_high(self):
        f = _make_finding(severity="High")
        assert f.target_closure_date(REF_DATE) == REF_DATE + __import__("datetime").timedelta(days=90)

    def test_target_closure_medium(self):
        f = _make_finding(severity="Medium")
        assert f.target_closure_date(REF_DATE) == REF_DATE + __import__("datetime").timedelta(days=180)

    def test_target_closure_low(self):
        f = _make_finding(severity="Low")
        assert f.target_closure_date(REF_DATE) == REF_DATE + __import__("datetime").timedelta(days=365)


class TestWriteFindings:
    def test_output_is_valid_xlsx(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        assert SHEET_NAME in wb.sheetnames

    def test_correct_number_of_data_rows(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        populated_rows = 0
        for row_num in range(DATA_START_ROW, DATA_START_ROW + 100):
            if ws.cell(row=row_num, column=DATA_START_COL).value is not None:
                populated_rows += 1
        assert populated_rows == len(findings)

    def test_finding_id_in_column_b(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        for i, f in enumerate(findings):
            cell_val = ws.cell(row=DATA_START_ROW + i, column=2).value
            assert cell_val == f.finding_id

    def test_severity_score_auto_derived(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        expected_scores = [4, 3, 2, 1]
        for i, expected in enumerate(expected_scores):
            cell_val = ws.cell(row=DATA_START_ROW + i, column=9).value  # col I
            assert cell_val == expected, f"Row {i}: expected score {expected}, got {cell_val}"

    def test_target_closure_dates(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        expected_dates = [
            REF_DATE + __import__("datetime").timedelta(days=90),   # Critical
            REF_DATE + __import__("datetime").timedelta(days=90),   # High
            REF_DATE + __import__("datetime").timedelta(days=180),  # Medium
            REF_DATE + __import__("datetime").timedelta(days=365),  # Low
        ]
        for i, expected in enumerate(expected_dates):
            cell_val = ws.cell(row=DATA_START_ROW + i, column=15).value  # col O
            if isinstance(cell_val, __import__("datetime").datetime):
                cell_val = cell_val.date()
            assert cell_val == expected, f"Row {i}: expected {expected}, got {cell_val}"

    def test_all_columns_populated(self, tmp_path: Path):
        findings = [_make_fixtures()[0]]
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        populated_cols = {
            2: "finding_id",      # B
            3: "audit_type",      # C
            4: "framework",       # D
            5: "audit_reference", # E
            6: "finding_description",  # F
            7: "finding_category",     # G
            8: "severity",        # H
            9: "severity_score",  # I
            10: "root_cause",     # J
            11: "affected_system",     # K
            12: "risk_owner",     # L
            13: "remediation_action",  # M
            14: "remediation_owner",   # N
            15: "target_closure_date", # O
            16: "current_status",      # P
            19: "evidence_reference",  # S
        }
        for col, field_name in populated_cols.items():
            val = ws.cell(row=DATA_START_ROW, column=col).value
            assert val is not None, f"Column {col} ({field_name}) should not be None"

    def test_header_row_preserved(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        assert ws.cell(row=8, column=2).value == "Finding ID"
        assert ws.cell(row=8, column=8).value == "Severity (Low/Medium/High/Critical)"

    def test_template_metadata_preserved(self, tmp_path: Path):
        out = write_findings([], tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        assert ws.cell(row=2, column=2).value == "Department:"

    def test_empty_findings_produces_valid_xlsx(self, tmp_path: Path):
        out = write_findings([], tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        assert SHEET_NAME in wb.sheetnames
        ws = wb[SHEET_NAME]
        assert ws.cell(row=DATA_START_ROW, column=2).value is None

    def test_single_finding(self, tmp_path: Path):
        findings = [_make_finding(finding_id="SOLO-1", severity="Medium")]
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        assert ws.cell(row=DATA_START_ROW, column=2).value == "SOLO-1"
        assert ws.cell(row=DATA_START_ROW, column=9).value == 2
        assert ws.cell(row=DATA_START_ROW + 1, column=2).value is None

    def test_output_creates_parent_dirs(self, tmp_path: Path):
        nested = tmp_path / "a" / "b" / "c" / "report.xlsx"
        write_findings([_make_finding()], nested, reference_date=REF_DATE)
        assert nested.exists()

    def test_framework_values(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        expected = ["SOC 2", "ISO 27001", "NIST CSF", "SOC 2"]
        for i, fw in enumerate(expected):
            assert ws.cell(row=DATA_START_ROW + i, column=4).value == fw

    def test_evidence_reference_populated(self, tmp_path: Path):
        findings = [_make_finding(evidence_reference="trail.jsonl#snap-42")]
        out = write_findings(findings, tmp_path / "report.xlsx", reference_date=REF_DATE)
        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]
        assert ws.cell(row=DATA_START_ROW, column=19).value == "trail.jsonl#snap-42"


class TestLoadFindingsFromJson:
    def test_load_list_format(self, tmp_path: Path):
        data = [
            {"finding_id": "F-1", "severity": "High"},
            {"finding_id": "F-2", "severity": "Low"},
        ]
        json_path = tmp_path / "evidence.json"
        json_path.write_text(__import__("json").dumps(data), encoding="utf-8")
        findings = load_findings_from_json(json_path)
        assert len(findings) == 2
        assert findings[0].finding_id == "F-1"
        assert findings[1].severity == "Low"

    def test_load_wrapped_format(self, tmp_path: Path):
        data = {"findings": [{"finding_id": "F-1", "severity": "Critical"}]}
        json_path = tmp_path / "evidence.json"
        json_path.write_text(__import__("json").dumps(data), encoding="utf-8")
        findings = load_findings_from_json(json_path)
        assert len(findings) == 1
        assert findings[0].severity_score == 4

    def test_defaults_applied(self, tmp_path: Path):
        data = [{"finding_id": "F-1"}]
        json_path = tmp_path / "evidence.json"
        json_path.write_text(__import__("json").dumps(data), encoding="utf-8")
        findings = load_findings_from_json(json_path)
        assert findings[0].audit_type == "Internal"
        assert findings[0].current_status == "Open"
        assert findings[0].severity == "Medium"


class TestGoldenFile:
    """End-to-end: write fixture findings, read back every cell, compare against expected values."""

    def test_golden_roundtrip(self, tmp_path: Path):
        findings = _make_fixtures()
        out = write_findings(findings, tmp_path / "golden.xlsx", reference_date=REF_DATE)

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb[SHEET_NAME]

        for i, f in enumerate(findings):
            row = DATA_START_ROW + i
            assert ws.cell(row=row, column=2).value == f.finding_id
            assert ws.cell(row=row, column=3).value == f.audit_type
            assert ws.cell(row=row, column=4).value == f.framework
            assert ws.cell(row=row, column=5).value == f.audit_reference
            assert ws.cell(row=row, column=6).value == f.finding_description
            assert ws.cell(row=row, column=7).value == f.finding_category
            assert ws.cell(row=row, column=8).value == f.severity
            assert ws.cell(row=row, column=9).value == f.severity_score
            assert ws.cell(row=row, column=10).value == f.root_cause
            assert ws.cell(row=row, column=11).value == f.affected_system
            assert ws.cell(row=row, column=12).value == f.risk_owner
            assert ws.cell(row=row, column=13).value == f.remediation_action
            assert ws.cell(row=row, column=14).value == f.remediation_owner

            expected_date = f.target_closure_date(REF_DATE)
            cell_val = ws.cell(row=row, column=15).value
            if isinstance(cell_val, __import__("datetime").datetime):
                cell_val = cell_val.date()
            assert cell_val == expected_date

            assert ws.cell(row=row, column=16).value == f.current_status
            assert ws.cell(row=row, column=19).value == f.evidence_reference

        next_row = DATA_START_ROW + len(findings)
        assert ws.cell(row=next_row, column=2).value is None
