"""Tests for multi-framework output — xlsx has N rows per framework reference."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from aec.formatter.audit_findings import (
    DATA_START_COL,
    DATA_START_ROW,
    SHEET_NAME,
    GapFinding,
    write_findings,
)
from aec.priors.framework_mapper import expand_findings_multi_framework

REF_DATE = date(2026, 5, 25)


def _base_finding(**overrides) -> GapFinding:
    defaults = {
        "finding_id": "AEC-CTRL-003-001",
        "audit_type": "Internal",
        "framework": "SOC 2",
        "audit_reference": "CC6.1",
        "finding_description": "Panel verdict: FAIL for CTRL-003",
        "finding_category": "Access Control",
        "severity": "High",
        "root_cause": "MFA not enforced",
        "current_status": "Open",
        "evidence_reference": "audit_trail.jsonl",
    }
    defaults.update(overrides)
    return GapFinding(**defaults)


class TestExpandFindingsMultiFramework:
    def test_single_finding_expands_to_three_frameworks(self):
        base = _base_finding()
        parsed_refs = [
            {"input": "SOC2:CC6.1", "display_fw": "SOC 2", "control_id": "CC6.1",
             "catalog_fw": "SOC 2", "category": "Access Control"},
            {"input": "ISO:A.8.2", "display_fw": "ISO 27001", "control_id": "A.8.2",
             "catalog_fw": "ISO 27001", "category": "Access Control"},
            {"input": "NIST-CSF:PR.AC-1", "display_fw": "NIST CSF", "control_id": "PR.AC-1",
             "catalog_fw": "NIST 800-53", "category": "Access Control"},
        ]
        expanded = expand_findings_multi_framework([base], parsed_refs)
        assert len(expanded) == 3

        frameworks = [f.framework for f in expanded]
        assert "SOC 2" in frameworks
        assert "ISO 27001" in frameworks
        assert "NIST CSF" in frameworks

        refs = [f.audit_reference for f in expanded]
        assert "CC6.1" in refs
        assert "A.8.2" in refs
        assert "PR.AC-1" in refs

    def test_finding_ids_unique(self):
        base = _base_finding()
        parsed_refs = [
            {"input": "SOC2:CC6.1", "display_fw": "SOC 2", "control_id": "CC6.1",
             "catalog_fw": "SOC 2", "category": "Access Control"},
            {"input": "ISO:A.8.2", "display_fw": "ISO 27001", "control_id": "A.8.2",
             "catalog_fw": "ISO 27001", "category": "Access Control"},
        ]
        expanded = expand_findings_multi_framework([base], parsed_refs)
        ids = [f.finding_id for f in expanded]
        assert len(ids) == len(set(ids))

    def test_evidence_reference_shared(self):
        base = _base_finding(evidence_reference="trail.jsonl#snap-003")
        parsed_refs = [
            {"input": "SOC2:CC6.1", "display_fw": "SOC 2", "control_id": "CC6.1",
             "catalog_fw": "SOC 2", "category": "Access Control"},
            {"input": "ISO:A.8.2", "display_fw": "ISO 27001", "control_id": "A.8.2",
             "catalog_fw": "ISO 27001", "category": "Access Control"},
        ]
        expanded = expand_findings_multi_framework([base], parsed_refs)
        for f in expanded:
            assert f.evidence_reference == "trail.jsonl#snap-003"

    def test_expansion_scopes_rows_to_internal_control_coverage(self):
        base = _base_finding(audit_reference="CTRL-007")
        parsed_refs = [
            {"input": "SOC2:CC6.1", "display_fw": "SOC 2", "control_id": "CC6.1",
             "catalog_fw": "SOC 2", "category": "Access Control"},
            {"input": "ISO:A.8.2", "display_fw": "ISO 27001", "control_id": "A.8.2",
             "catalog_fw": "ISO 27001", "category": "Access Control"},
            {"input": "NIST-CSF:PR.AC-1", "display_fw": "NIST CSF", "control_id": "PR.AC-1",
             "catalog_fw": "NIST CSF", "category": "Access Control"},
        ]
        coverage = {
            "SOC2:CC6.1": ["CTRL-002", "CTRL-003"],
            "ISO:A.8.2": ["CTRL-003", "CTRL-007"],
            "NIST-CSF:PR.AC-1": ["CTRL-002", "CTRL-003", "CTRL-007"],
        }
        expanded = expand_findings_multi_framework([base], parsed_refs, coverage)
        assert [f.framework for f in expanded] == ["ISO 27001", "NIST CSF"]


class TestMultiFrameworkXlsx:
    def test_xlsx_has_n_rows_per_framework(self, tmp_path: Path):
        base = _base_finding()
        parsed_refs = [
            {"input": "SOC2:CC6.1", "display_fw": "SOC 2", "control_id": "CC6.1",
             "catalog_fw": "SOC 2", "category": "Access Control"},
            {"input": "ISO:A.8.2", "display_fw": "ISO 27001", "control_id": "A.8.2",
             "catalog_fw": "ISO 27001", "category": "Access Control"},
            {"input": "NIST-CSF:PR.AC-1", "display_fw": "NIST CSF", "control_id": "PR.AC-1",
             "catalog_fw": "NIST 800-53", "category": "Access Control"},
        ]
        expanded = expand_findings_multi_framework([base], parsed_refs)
        out = write_findings(expanded, tmp_path / "report.xlsx", reference_date=REF_DATE)

        wb = openpyxl.load_workbook(out)
        ws = wb[SHEET_NAME]

        populated = 0
        for row in range(DATA_START_ROW, DATA_START_ROW + 20):
            if ws.cell(row=row, column=DATA_START_COL).value is not None:
                populated += 1
        assert populated == 3

        frameworks_in_xlsx = []
        for row in range(DATA_START_ROW, DATA_START_ROW + 3):
            frameworks_in_xlsx.append(ws.cell(row=row, column=4).value)
        assert set(frameworks_in_xlsx) == {"SOC 2", "ISO 27001", "NIST CSF"}
